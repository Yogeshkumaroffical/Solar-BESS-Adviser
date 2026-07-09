"""
Export service: generates PDF (WeasyPrint) and Excel (openpyxl) reports.
"""

from __future__ import annotations
import io
from datetime import datetime
from models.outputs import AnalysisResult


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------

def generate_excel(result: AnalysisResult, project_name: str = "Solar BESS Project") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    
    # Color palette
    HEADER_FILL = PatternFill("solid", fgColor="1A2744")
    SOLAR_FILL = PatternFill("solid", fgColor="F59E0B")
    HYBRID_FILL = PatternFill("solid", fgColor="10B981")
    ALT_FILL = PatternFill("solid", fgColor="F8FAFC")
    
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(cell, color_fill=HEADER_FILL):
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = color_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    def style_data(cell, alt=False):
        cell.alignment = Alignment(horizontal="right")
        cell.border = border
        if alt:
            cell.fill = ALT_FILL

    # ------ Sheet 1: Summary ------
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 20
    ws1.column_dimensions["D"].width = 20

    ws1["A1"] = f"Solar BESS Profitability Report: {project_name}"
    ws1["A1"].font = Font(bold=True, size=14, color="1A2744")
    ws1["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws1["A2"].font = Font(italic=True, color="64748B")
    ws1["A3"] = f"Recommendation: {result.recommendation}"
    ws1["A3"].font = Font(bold=True, size=12, color="059669" if result.recommendation == "Attractive" else "DC2626")

    headers = ["Metric", "Solar Only", "Solar + BESS", "Incremental"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=5, column=col, value=h)
        style_header(cell)

    s = result.solar_only_financials
    h = result.hybrid_financials
    rows = [
        ("NPV (₹ Cr)", s.npv_cr, h.npv_cr, result.incremental_npv_cr),
        ("IRR (%)", s.irr_percent, h.irr_percent, "—"),
        ("CAPEX (₹ Cr)", s.total_capex_cr, h.total_capex_cr, h.total_capex_cr - s.total_capex_cr),
        ("BESS Payback (yrs)", "—", h.simple_payback_years, "—"),
        ("LCOE (₹/kWh)", s.lcoe_inr_kwh, h.lcoe_inr_kwh, "—"),
    ]
    for r_idx, row in enumerate(rows, 6):
        for c_idx, val in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            if c_idx == 1:
                cell.font = Font(bold=True)
            else:
                cell.number_format = "#,##0.00"
            style_data(cell, alt=r_idx % 2 == 0)

    # ------ Sheet 2: Annual Cash Flows ------
    ws2 = wb.create_sheet("Annual Cash Flows")
    for col_idx, w in enumerate([5, 12, 18, 18, 18, 18, 18, 18], 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = w

    cf_headers = ["Year", "Generation (MWh)", "Solar Rev (₹Cr)", "Solar Profit (₹Cr)",
                  "Solar Cum CF (₹Cr)", "Hybrid Rev (₹Cr)", "Hybrid Profit (₹Cr)", "Hybrid Cum CF (₹Cr)"]
    for col, h in enumerate(cf_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        style_header(cell)

    for r_idx, cf in enumerate(result.annual_cash_flows, 2):
        vals = [cf.year, round(cf.generation_mwh, 0), round(cf.solar_only_revenue_cr, 2),
                round(cf.solar_only_profit_cr, 2), round(cf.solar_only_cumulative_cf_cr, 2),
                round(cf.hybrid_revenue_cr, 2), round(cf.hybrid_profit_cr, 2), round(cf.hybrid_cumulative_cf_cr, 2)]
        for c_idx, val in enumerate(vals, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.number_format = "#,##0.00" if c_idx > 1 else "0"
            style_data(cell, alt=r_idx % 2 == 0)

    # ------ Sheet 3: Revenue Breakdown Y1 ------
    ws3 = wb.create_sheet("Year-1 Revenue Breakdown")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 20

    for col, h in enumerate(["Revenue Stream", "Solar Only (₹Cr)", "Hybrid (₹Cr)"], 1):
        cell = ws3.cell(row=1, column=col, value=h)
        style_header(cell)

    so = result.solar_only_year1
    hy = result.hybrid_year1
    bd_rows = [
        ("Base PPA Revenue", so.base_ppa_revenue_cr, hy.base_ppa_revenue_cr),
        ("DSM Penalty", so.dsm_penalty_cr, hy.dsm_penalty_cr),
        ("DSM Savings", 0, hy.dsm_savings_cr),
        ("Clipping Recovery", 0, hy.clipping_recovery_cr),
        ("Curtailment Recovery", 0, hy.curtailment_recovery_cr),
        ("Arbitrage", 0, hy.arbitrage_cr),
        ("Ancillary Services", 0, hy.ancillary_cr),
        ("Total O&M", -so.total_om_cr, -hy.total_om_cr),
        ("Net Profit (pre-tax)", so.net_profit_cr, hy.net_profit_cr),
        ("Net Profit (after-tax)", so.net_profit_after_tax_cr, hy.net_profit_after_tax_cr),
    ]
    for r_idx, row in enumerate(bd_rows, 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)
            if c_idx == 1:
                cell.font = Font(bold=(r_idx in [9, 10]))
            else:
                cell.number_format = "#,##0.00"
            style_data(cell, alt=r_idx % 2 == 0)

    # ------ Sheet 4: Sensitivity ------
    ws4 = wb.create_sheet("Sensitivity")
    ws4.column_dimensions["A"].width = 25
    ws4.column_dimensions["B"].width = 15
    ws4.column_dimensions["C"].width = 20
    ws4.column_dimensions["D"].width = 20
    ws4.column_dimensions["E"].width = 20

    for col, h in enumerate(["Variable", "Change (%)", "Solar NPV (₹Cr)", "Hybrid NPV (₹Cr)", "NPV Gain (₹Cr)"], 1):
        cell = ws4.cell(row=1, column=col, value=h)
        style_header(cell)

    for r_idx, sp in enumerate(result.sensitivity, 2):
        vals = [sp.variable, sp.change_percent, sp.solar_only_npv_cr, sp.hybrid_npv_cr, sp.npv_gain_cr]
        for c_idx, val in enumerate(vals, 1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=val)
            cell.number_format = "#,##0.00" if c_idx > 1 else "@"
            style_data(cell, alt=r_idx % 2 == 0)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF Export (WeasyPrint via HTML template)
# ---------------------------------------------------------------------------

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap');
  body {{ font-family: 'Inter', Arial, sans-serif; color: #1e293b; margin: 0; padding: 24px; font-size: 11px; }}
  h1 {{ color: #1a2744; font-size: 20px; margin-bottom: 4px; }}
  h2 {{ color: #1a2744; font-size: 14px; border-bottom: 2px solid #e2e8f0; padding-bottom: 4px; margin-top: 20px; }}
  .subtitle {{ color: #64748b; font-size: 10px; margin-bottom: 16px; }}
  .recommendation {{ padding: 10px 16px; border-radius: 6px; font-size: 13px; font-weight: 700; margin: 12px 0; }}
  .Attractive {{ background: #d1fae5; color: #065f46; }}
  .Marginal {{ background: #fef3c7; color: #92400e; }}
  .Not-Attractive {{ background: #fee2e2; color: #991b1b; }}
  table {{ width: 100%; border-collapse: collapse; margin: 12px 0; font-size: 10px; }}
  th {{ background: #1a2744; color: white; padding: 7px 10px; text-align: left; }}
  td {{ padding: 6px 10px; border-bottom: 1px solid #e2e8f0; }}
  tr:nth-child(even) td {{ background: #f8fafc; }}
  .num {{ text-align: right; font-variant-numeric: tabular-nums; }}
  .footer {{ margin-top: 40px; color: #94a3b8; font-size: 9px; text-align: center; }}
  .kpi-grid {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 10px; margin: 12px 0; }}
  .kpi {{ background: #f1f5f9; border-radius: 6px; padding: 10px; text-align: center; }}
  .kpi-val {{ font-size: 20px; font-weight: 700; color: #1a2744; }}
  .kpi-label {{ font-size: 9px; color: #64748b; margin-top: 2px; }}
</style>
</head>
<body>
<h1>Solar + BESS Profitability Report</h1>
<div class="subtitle">Project: {project_name} &nbsp;|&nbsp; Generated: {date}</div>

<div class="recommendation {rec_class}">{recommendation}: {recommendation_reason}</div>

<div class="kpi-grid">
  <div class="kpi"><div class="kpi-val">₹{hybrid_npv:.1f} Cr</div><div class="kpi-label">Hybrid NPV</div></div>
  <div class="kpi"><div class="kpi-val">{hybrid_irr}%</div><div class="kpi-label">Hybrid IRR</div></div>
  <div class="kpi"><div class="kpi-val">{payback} yrs</div><div class="kpi-label">BESS Payback</div></div>
</div>

<h2>Financial Summary</h2>
<table>
<tr><th>Metric</th><th class="num">Solar Only</th><th class="num">Solar + BESS</th><th class="num">Incremental</th></tr>
<tr><td>NPV (₹ Cr)</td><td class="num">{solar_npv:.1f}</td><td class="num">{hybrid_npv:.1f}</td><td class="num">{incr_npv:.1f}</td></tr>
<tr><td>IRR (%)</td><td class="num">{solar_irr}</td><td class="num">{hybrid_irr}</td><td class="num">—</td></tr>
<tr><td>CAPEX (₹ Cr)</td><td class="num">{solar_capex:.1f}</td><td class="num">{hybrid_capex:.1f}</td><td class="num">{bess_capex:.1f}</td></tr>
<tr><td>BESS Payback (yrs)</td><td class="num">—</td><td class="num">{payback}</td><td class="num">—</td></tr>
</table>

<h2>Year-1 Revenue Breakdown</h2>
<table>
<tr><th>Stream</th><th class="num">Solar Only (₹ Cr)</th><th class="num">Hybrid (₹ Cr)</th></tr>
{revenue_rows}
</table>

<h2>Top Sensitivity Drivers (Hybrid NPV Swing)</h2>
<table>
<tr><th>Variable</th><th class="num">Change</th><th class="num">Hybrid NPV (₹ Cr)</th><th class="num">NPV Gain (₹ Cr)</th></tr>
{sensitivity_rows}
</table>

<div class="footer">
  Solar BESS Profitability Advisor &nbsp;|&nbsp; All values in INR Crores &nbsp;|&nbsp; For advisory purposes only
</div>
</body>
</html>"""


def generate_pdf(result: AnalysisResult, project_name: str = "Solar BESS Project") -> bytes:
    try:
        from weasyprint import HTML as WeasyHTML
    except ImportError:
        # WeasyPrint requires system GTK/Cairo libraries.
        # Fall back to a simple HTML-as-PDF approach or raise a clear error.
        raise RuntimeError(
            "PDF export requires WeasyPrint which needs GTK/Cairo system libraries. "
            "Install WeasyPrint (pip install weasyprint) and its system dependencies, "
            "or use the Excel export instead."
        )

    s = result.solar_only_financials
    h = result.hybrid_financials
    so = result.solar_only_year1
    hy = result.hybrid_year1

    revenue_stream_rows = [
        ("Base PPA Revenue", so.base_ppa_revenue_cr, hy.base_ppa_revenue_cr),
        ("DSM Penalty", so.dsm_penalty_cr, hy.dsm_penalty_cr),
        ("DSM Savings", 0, hy.dsm_savings_cr),
        ("Clipping Recovery", 0, hy.clipping_recovery_cr),
        ("Curtailment Recovery", 0, hy.curtailment_recovery_cr),
        ("Arbitrage", 0, hy.arbitrage_cr),
        ("Ancillary Services", 0, hy.ancillary_cr),
        ("Total O&M", -so.total_om_cr, -hy.total_om_cr),
        ("Net Profit (after-tax)", so.net_profit_after_tax_cr, hy.net_profit_after_tax_cr),
    ]
    revenue_rows = "\n".join(
        f"<tr><td>{n}</td><td class='num'>{sv:.2f}</td><td class='num'>{hv:.2f}</td></tr>"
        for n, sv, hv in revenue_stream_rows
    )

    # Top 8 sensitivity points by absolute NPV swing
    top_sens = sorted(result.sensitivity, key=lambda x: abs(x.npv_gain_cr - result.incremental_npv_cr), reverse=True)[:8]
    sensitivity_rows = "\n".join(
        f"<tr><td>{sp.variable}</td><td class='num'>{sp.change_percent:+.0f}%</td><td class='num'>{sp.hybrid_npv_cr:.1f}</td><td class='num'>{sp.npv_gain_cr:.1f}</td></tr>"
        for sp in top_sens
    )

    rec_class = result.recommendation.replace(" ", "-")
    html_content = HTML_TEMPLATE.format(
        project_name=project_name,
        date=datetime.now().strftime("%d %b %Y"),
        recommendation=result.recommendation,
        recommendation_reason=result.recommendation_reason[:120] + "..." if len(result.recommendation_reason) > 120 else result.recommendation_reason,
        rec_class=rec_class,
        hybrid_npv=h.npv_cr,
        hybrid_irr=f"{h.irr_percent:.1f}" if h.irr_percent else "N/A",
        solar_irr=f"{s.irr_percent:.1f}" if s.irr_percent else "N/A",
        payback=f"{h.simple_payback_years:.1f}" if h.simple_payback_years else "N/A",
        solar_npv=s.npv_cr,
        incr_npv=result.incremental_npv_cr,
        solar_capex=s.total_capex_cr,
        hybrid_capex=h.total_capex_cr,
        bess_capex=h.total_capex_cr - s.total_capex_cr,
        revenue_rows=revenue_rows,
        sensitivity_rows=sensitivity_rows,
    )

    pdf_bytes = WeasyHTML(string=html_content).write_pdf()
    return pdf_bytes

